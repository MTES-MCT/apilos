import json

from io import BytesIO
from openpyxl import load_workbook
from django.conf import settings


def assert_xlsx(self, my_class, file_name):
    filepath = f"{settings.BASE_DIR}/static/files/{file_name}.xlsx"
    with open(filepath, "rb") as excel:
        my_wb = load_workbook(filename=BytesIO(excel.read()), data_only=True)
        self.assertIn(my_class.sheet_name, my_wb)
        my_ws = my_wb[my_class.sheet_name]

        column_values = []
        for col in my_ws.iter_cols(
            min_col=1, max_col=my_ws.max_column, min_row=1, max_row=1
        ):
            for cell in col:
                column_values.append(cell.value)

        for key in my_class.import_mapping:
            words = key.split()
            exists_in_list = False
            for value in column_values:
                if value is not None and all(word in value for word in words):
                    exists_in_list = True
            self.assertTrue(exists_in_list, f"not found ${key} in column ")


def assert_get_text_and_files(self, object_to_test, field) -> None:
    assert_get_text(self, object_to_test, field)
    assert_get_files(self, object_to_test, field)


def assert_get_text(self, object_to_test, field) -> None:
    func_text = getattr(object_to_test, field + "_text")
    text = func_text()
    try:
        expected_text = json.loads(getattr(object_to_test, field))["text"]
    except TypeError:
        expected_text = ""
    except json.decoder.JSONDecodeError:
        expected_text = ""
    self.assertEqual(text, expected_text)


def assert_get_files(self, object_to_test, field) -> None:
    func_files = getattr(object_to_test, field + "_files")
    files = func_files()
    try:
        expected_files = json.loads(getattr(object_to_test, field))["files"]
    except TypeError:
        expected_files = {}
    except json.decoder.JSONDecodeError:
        expected_files = {}
    self.assertEqual(files, expected_files)
