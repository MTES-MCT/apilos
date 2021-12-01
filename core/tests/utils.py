import datetime
import json
import random

from io import BytesIO
from openpyxl import load_workbook
from django.conf import settings
from bailleurs.models import Bailleur
from programmes.models import Programme


def create_bailleur():
    return Bailleur.objects.create(
        nom="3F",
        siret="12345678901234",
        capital_social="123000.50",
        ville="Marseille",
        signataire_nom="Patrick Patoulachi",
        signataire_fonction="PDG",
        signataire_date_deliberation=datetime.date(2014, 10, 9),
    )


def create_programme(bailleur, administration=None):
    files_and_text = (
        '{"files": {"bbfc7e3a-e0e7-4899-a1e1-fc632c3ea6b0": {"uuid": "bbfc7e3a'
        + '-e0e7-4899-a1e1-fc632c3ea6b0", "thumbnail": "data:image/png;base64,'
        + "iVBORw0KGgoAAAANSUhEUgAAAHgAAAB4CAYAAAA5ZDbSAAAAAXNSR0IArs4c6QAAB2Z"
        + "JREFUeF7tmn1I1VcYx79XN99yN0vSmE6zCBqDCkV6EXWLomWskeH8w4yaOMIpNVDSJm"
        + "UFvZBEoVbMZk2LsKA25rJXxnoZslQqJhPLGDWlNIc5UxvqHefEvfhSDJ+fv3t7bs8Bu"
        + "fzu7/d9zvd8P/ccr/JYnvf12CDDbROwCGC3ZasXJoDdm68AdnO+AlgAu3sCbr4++R0s"
        + "gN08ATdf3it38COrH6Z29bj58t1/eaMA2zqe4HFE2Iu/od72gn9uLibk5L4yiYFBGxp"
        + "+/xvTpr0Dq9XL/RNjtkJLVFSUrb29HXebGodYt+GRdQKmVJ+HZ0zcK5eUnV2DRYtCkJ"
        + "DwHh4+fIYPY79H858pzCJwb7uW5M+SbOXl3w1b5eOQIAS3tOH/jumCgjoUFEThk4Rz+"
        + "PFcAp486UN/vw1Tp/q6d2qMVme5dave9v6sWWO2/PTpvxgYsGHyZG9YfUrR1Zeua5QU"
        + "N+DLzA/GXE8E5iRABqzspH3+C74ti3cAbml5Bj+/tzBpkrc5bqXqmBMwBLisrAl/Pfw"
        + "HX+dHora2HV9lXcevvyWO2YQIzEtgXP7R8eBBN0JCJsDT02KeU6lMSmBcAJNmFpFTEh"
        + "DATonZdZMIYNdl75SZBbBTYnbdJALYddk7ZWYB7JSYXTeJAHZd9k6ZWQA7JWbXTWLp7"
        + "OyUvmjX5W/6zALY9IhdO4EAdm3+ps8ugE2P2LUTCGDX5m/67ALY9IhdO8G4AF67di2O"
        + "HTvm2pXI7C9NYFwAt7W1ISgoSCJ+DRMYBnj27Nnw9fXFtWvX4OXlBXXt4+OjX48cOaJ"
        + "f79y5M2wZ8fHx6OzsxO3bt3H16lVkZ2frn5KSEqh7Fy9eRE1NDbKyslBUVIT9+/dj48"
        + "aNiIuL07qR9V7DjFhbGrWDFSgF2MPDAxkZGXpxO3fuxObNm0cBjomJ0aDUCAgIwIIFC"
        + "xzXKSkpCAsLQ0tLC8rLy18KWH0gZJibwDDAXV1dsFqtKC4uRnBwMJKSkvTsCq6CbN/B"
        + "HR0dCAwMxOLFi3H58mWHQwXMDjwvLw+7du1y3EtOTkZlZSVycnKwd+9e/ZwANheuqj7"
        + "qiFYgJk6ciMzMTKgd2t3djVWrVmHLli3o7e3FvHnzNOjjx49rd3PnzsXg4CBqa2v1UW"
        + "wHrHbu8uXLERoaiqqqKn0/LS0NO3bswIoVKwSw+Wz1DA7AfhXf4PlHH8P75/PoSf0C6"
        + "now+F1Yentw/acqRBcWweeHSsc9W+CUFwU62h3v9X2arJ+xWQNg8/WDx+NWxz17bfj6"
        + "YSA0HJ53/9D3ZJibwLh8izbXolQ3koAANpIeA60AZgDJiEUBbCQ9BloBzACSEYsC2Eh"
        + "6DLQCmAEkIxYFsJH0GGgFMANIRiwKYCPpMdAKYAaQjFgUwEbSY6AVwAwgGbEogI2kx0"
        + "BrGuCoqCjU1dUZjuDKlSu6M2TmzJmGa72JBUwDPB5hzpkzR/d63bx5E9HR0eNR8o2rY"
        + "Zk+fbqtvr7e0Y6jGuFUP9bp06dRVlam+6lUf1V+fr5uwblx4wZOnDih31Nj5cqVOHv2"
        + "LBQMdV89N7SePdGFCxdi06ZNuvFOtflUVFTA29sbBw4c0DVjY2Oxe/du/bjqJBk6jh4"
        + "9inXr1r1xcMZjwXoH9/f3IzIyUnc4DgWsWmzUGLmDRgI+efKk7r60j6H17O+lp6ejtL"
        + "RUXyYmJuLMmTO4f/8+Dh48iMLCQmzbtg0RERFYs2bNsHU1NzejsbFRt//IGHsClry8P"
        + "JvaWfaGunv37mHGjBl699oBp6am6h2n4KvuSfVBUM+rsWHDBr0LVT/W/PnzsWfPHr1T"
        + "R7bYLlu2DNXV1VqjgF66dAkXLlzQPV/+/v4O51u3btWw1WhtbUVDQwOWLFky9pWJQid"
        + "gaWpqstmb1levXo2lS5dCAVXHpv2oVN2W6ojct28fwsPDtTA3N1cDP3z4sGOXqyP20K"
        + "FDugl++/btulFv6A5WR7/Sqy5NtcvXr1+v+63tp4Tqtjx16pRDo2rYh9rB6oubjLEl4"
        + "LQvWUOP6LFZlKeNJOA0wEZMipaegACmZ8dCKYBZYKKbFMD07FgoBTALTHSTApieHQul"
        + "AGaBiW5SANOzY6EUwCww0U0KYHp2LJQCmAUmukkBTM+OhVIAs8BENymA6dmxUApgFpj"
        + "oJgUwPTsWSgHMAhPdpACmZ8dCKYBZYKKbFMD07FgoBTALTHSTApieHQulAGaBiW5SAN"
        + "OzY6EUwCww0U0KYHp2LJQCmAUmukkBTM+OhVIAs8BENymA6dmxUApgFpjoJgUwPTsWS"
        + "gHMAhPdpACmZ8dCKYBZYKKbFMD07FgoBTALTHSTApieHQulAGaBiW5SANOzY6EUwCww"
        + "0U0KYHp2LJQCmAUmukkBTM+OhVIAs8BENymA6dmxUApgFpjoJgUwPTsWSgHMAhPdpAC"
        + "mZ8dCKYBZYKKbFMD07FgoBTALTHSTApieHQulAGaBiW5SANOzY6G0ZGRk2Fg4FZOkBC"
        + "zP+3oEMCk6HiIBzIMT2aUAJkfHQyiAeXAiuxTA5Oh4CP8DBymut3Zy8jIAAAAASUVOR"
        + 'K5CYII=", "size": 31185, "filename": "acquereur1.png"}, "9e69e766-0'
        + '167-4638-b1ce-f8f0b033e03a": {"uuid": "9e69e766-0167-4638-b1ce-f8f0'
        + 'b033e03a", "thumbnail": "data:image/png;base64,iVBORw0KGgoAAAANSUhE'
        + "UgAAAHgAAAB4CAYAAAA5ZDbSAAAAAXNSR0IArs4c6QAACJ1JREFUeF7tnX1MVWUcx7+"
        + "XlwvCLl5QQI0QDEh8WQxWJhMXBGZNnRr/iEmzuUVW5uZWppJrutlcyyRntWq5NeGf3F"
        + "DbNDWZbY6sKAmF4sUrZCSIdBGkC9wL7XmO98ybgg/3Xj33Oed3/rn3Oed5/X1+L+d57"
        + "j3PMQ04+kdAh24lYCLAumXLB0aA9c2XAOucLwH2B+Co8M/8Uc19qYNctB/ESoD9IMRA"
        + "roIABzIdP/SNAPtBiIFcBQEOZDp+6BsB9oMQA7mK8QI+d361x3BeWvMtLtZ1iQ3RBER"
        + "EhqK/bwiJ0y1oa+3l5WakWGFr6cHIiOfCJN1F30OsLpcLwcHBY+YaL+D/VxYbH4FrHf"
        + "0ep9PnxiJ9ziSUly9CUdEJnPjmEvp6B3HT+ToiQz7iefMKEnH6ZBv/fvL7QhQs/PqOf"
        + "hLgu6Cz2+2wWq38Snl5BYqKVvkdsGNkA4aHRxARrMAa7WD5wk1l/PLf3S+jq8uBNUXH"
        + "cf7nDhVw7+Br/Hxa8pdotK1F8rTPef7fbWtpoYMJ4pnFz2H9KyWYOXMmSt/ZjpCQEJQ"
        + "f/ArnfvwJ8554HE/nL8LCnAUoWFSAJ+fNQ1BQEBdgZ2cn4uLi4I0Fh4QEYdrDFrTZeo"
        + "QAP7c8Bde7/+V5z5x5nkNnFpyTm4jjx23q+ZioT+AcdCFt1iRc+LWTAA8ODmLFykLF5"
        + "eU+heTkZKxcuYKnGdgXXyxG8ZoXeDq/4Bnk5eViy9ub4XQ6uSKw40EAzluchNPHL3so"
        + "AwP8RPY0vPfuDx7nexyvIijIBIt5HwEeHh7G+fO1yMh4DL29vdi8eQv279/Hv0dFRaG"
        + "m5hekpDyCgYEBXLh4EZWVh1G290Pk5uWj6vQprwGPaba3XbzdRTe2roXVGoaKiia8Uf"
        + "Kd6qI7e0pgNgfhrTer8elHv6J0Zzba/+rDFx//RoCZLBlkdjMVGhrKRetyDWPxs8/i5"
        + "IlveXpoaIjfaLldM4NvsVhUDN5YsChgX/PRTdYoEuzo6EB8fLyQfAmwkJjkzUSA5WUn"
        + "1HMCLCQmeTMRYHnZCfX86BHP6YtQoQeUiW6yHpCgtWrGZLfbdfe32QnhZq3kGXDt6hI"
        + "wkzJBVnSNAN+yOYdjAE3NLXz5kS1DsiMrK4t/1tTUjJpmS511dXXq9a6uLrS2tqppm8"
        + "2G7u5uNV1bW+tRv7tud3v+ThNgACGhZly6ZONwp0+fDpPJFHCu1tsOEWAAjU0tmDNnj"
        + "rcyDOhyhgf855W/YLf3qO44oGl50TnDA667UO8RX72QYUAXMTxgc9iEgAbka+cMD5gs"
        + "2FcV0qi86DyYAGsEyNdmCTAtdHAJUAz21ZQ0Kk8WTBbMJUAxWCML9LVZsmCyYIrBvlq"
        + "RluXJgsmCKQZraYG+ti1qwRcuNvCmMjMzfW0yIMsbfqmS5sEBqZf37pSoBdM06d6yDM"
        + "gcBJhusrgELtmU/0+xZ4P1eFAMpt+D5dRrctHkomkeLKftKr0WteCrHdd4/sTERJmHO"
        + "2rfKQZTDJZTsUUtmObBcvIVdtFaAWaPs6xevZpPz/bs2cOl3NDQwPcLSUlJQVhYmF8k"
        + "b3gXfaP3Jhfk5MmT/SJQkUrYJi4HDx5ESUkJz15UVITy8nIOOD09XaQK4TyGB6zFWnR"
        + "ubi6qqqrugMQAs2ejIiIicOzYMf6sVGpqKlatWoXq6mosW7YMR44cQX19PWbNmoUNGz"
        + "Zg7969OHz4MJYsWYKmpiZe57p163D27Fn+3fCAtXDRS5cuxdGjR8cEPH/+fA6VHe7v/"
        + "wfMzpvNyrPQu3btQnZ2NvLz83HqlLJ/FwHW6D9ZbEfYwsJCHDp0iMfcnJwcbnG3W/C2"
        + "bduwY8cO/qSjG7D7c+fOnWDXly9fjsrKSg6SbdS2detW7N69W93PiwAD6P5H2StyypQ"
        + "pwnHNXxkrKiowd+7cUZ9svHLlCtjzx+xmzG3NBw4cQHFxsQqxubkZV69exYIFC+7aLc"
        + "O7aC1i8HgV5HZ3Pd6yhgesRQweLyRf8hNgenzUF/3RrqzoStbNfgfvZHR0tHadvY8tG"
        + "96CZYjBvvA3PGCKwb6oj4ZlRV00AdYQki9NiwJ2upSN/tjyoB4Pw7toU5Dy3gX3bu96"
        + "g2x4wOSiJVVpURdNgHUOODhE+TXmXm83k1QM9HPhkHOYs4uMjJSV4Zj9phhMS5VyKjb"
        + "FYIWb4S2YlirlNGDhf1X23VRe+BgTEyPpSMfutuEtmKZJkuo1xWCKwVwCFIN1bsHXu+"
        + "18hFOnTpV0pBSDx5QAxWBJ9ZpiMMVgisGSGi/vtqgFt//dwfMnJSXJPNxR+07zYFqLl"
        + "lOxRS2YbrLk5CvsomkerHPAl1v/5CNMS0uTdKQ0D6Z5sB5Vl2IwzYNpHiyzZYta8B+N"
        + "zXyY7EFsPR40D6Z5sJx6LWrBNA+Wky/Ng29xIxdNLlpOEyYXTdMkLgGKwXIaMMVgisG"
        + "KBMiCdW7BBJgASyoBusmitWiZVZemSWTBdJNFFiyzBMiCKQbLrL8Ug8mCKQYbwYLrG3"
        + "6HyzWMrKwsmYc7at8N/3Mhe+lFqDlcl3DZoAwP2L0ezTZCy8jI0B1o3QEuKyvDxo0bE"
        + "WZWNhkVPdiaNAPc2dmJ9vZ2Xoy5bfaqm/7+fjVdW1sLp9OppmtqatQmWH7RNPMcmZmZ"
        + "av7w8HDMnj1bTU+cOJG/4s5dX2xsLH9DqjvNdqifMWOGmk5ISEB8fLyaZn/kZzvo6go"
        + "wg7tp0yb09fXBOtEiylbNx4R+7dp13Oi9wc+lpT2KtrZWOBzKtv8s3dLSDJfLpaYbG/"
        + "9Qy7PromnWVmpqmpqfveAqKSlZTbOd9x56KEFNW61WxMXFq2mLJYrvSuBujylAdHSMm"
        + "k5IeBhMaXQDmL35q7S0FNu3b+efotOkcWuBZAVM69evV3bE1tmx54P3dTYi74ZjGnD0"
        + "6xKwd+LQXykCrD+mHiMiwARY5xLQ+fDIgnUO+D8ei8UeY8TTjQAAAABJRU5ErkJggg="
        + '=", "size": 69076, "filename": "acquereur2.png"}}, "text": "yala !!'
        + '!!"}'
    )

    return Programme.objects.create(
        nom="3F",
        administration=administration,
        bailleur=bailleur,
        code_postal="75007",
        ville="Paris",
        adresse="22 rue segur",
        departement=75,
        numero_galion="12345",
        annee_gestion_programmation=2018,
        zone_123=3,
        zone_abc="B1",
        surface_utile_totale=5243.21,
        nb_locaux_commerciaux=5,
        nb_bureaux=25,
        autre_locaux_hors_convention="quelques uns",
        vendeur=random.choice([files_and_text, "", "n'importe quoi", None]),
        acquereur=random.choice([files_and_text, "", "n'importe quoi", None]),
        reference_notaire=random.choice([files_and_text, "", "n'importe quoi", None]),
        reference_publication_acte=random.choice(
            [files_and_text, "", "n'importe quoi", None]
        ),
        acte_de_propriete=random.choice([files_and_text, "", "n'importe quoi", None]),
        acte_notarial=random.choice([files_and_text, "", "n'importe quoi", None]),
        reference_cadastrale=random.choice(
            [files_and_text, "", "n'importe quoi", None]
        ),
        edd_volumetrique=random.choice([files_and_text, "", "n'importe quoi", None]),
        edd_classique=random.choice([files_and_text, "", "n'importe quoi", None]),
        permis_construire="123 456 789 ABC",
        date_achevement_previsible=datetime.date.today() + datetime.timedelta(days=365),
        date_achat=datetime.date.today() - datetime.timedelta(days=365),
        date_achevement=datetime.date.today() + datetime.timedelta(days=465),
    )


def assert_xlsx(self, my_class, file_name):
    filepath = f"{settings.BASE_DIR}/static/files/{file_name}.xlsx"
    with open(filepath, "rb") as excel:
        my_wb = load_workbook(filename=BytesIO(excel.read()), data_only=True)
        self.assertIn(my_class.sheet_name, my_wb)
        my_ws = my_wb[my_class.sheet_name]

        column_values = []
        for col in my_ws.iter_cols(
            min_col=1, max_col=my_ws.max_column, min_row=1, max_row=1
        ):
            for cell in col:
                column_values.append(cell.value)

        for key in my_class.import_mapping:
            self.assertIn(key, column_values)


def assert_get_text_and_files(self, object_to_test, field) -> None:
    assert_get_text(self, object_to_test, field)
    assert_get_files(self, object_to_test, field)


def assert_get_text(self, object_to_test, field) -> None:
    func_text = getattr(object_to_test, field + "_text")
    text = func_text()
    try:
        expected_text = json.loads(getattr(object_to_test, field))["text"]
    except TypeError:
        expected_text = ""
    except json.decoder.JSONDecodeError:
        expected_text = ""
    self.assertEqual(text, expected_text)


def assert_get_files(self, object_to_test, field) -> None:
    func_files = getattr(object_to_test, field + "_files")
    files = func_files()
    try:
        expected_files = json.loads(getattr(object_to_test, field))["files"]
    except TypeError:
        expected_files = {}
    except json.decoder.JSONDecodeError:
        expected_files = {}
    self.assertEqual(files, expected_files)
