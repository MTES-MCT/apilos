import os
import sys

from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from instructeurs.models import Administration
from bailleurs.models import Bailleur
from programmes.models import Programme, Financement, Lot, TypeStationnement


def str_row(row):
    return (
        f"Bailleur (siret, nom) : {row['MOA (code SIRET)']} - {row['MOA (nom officiel)']}"
        + f", Administration (code, nom) : {row['Gestionnaire (code)']} - {row['Gestionnaire']}"
        + f", Programme (nom, ville) : {row['Nom Opération']} - {row['Commune']},"
        + f" Financement : {row['Produit']}"
    )


class Command(BaseCommand):
    # pylint: disable=R0912,R0914,R0915
    def handle(self, *args, **options):
        basedir = settings.BASE_DIR
        file_path = os.path.join(basedir, "documents", "v3.xlsx")

        file_path_input = input("Enter your value (default: " + file_path + "): ")
        file_path = file_path_input or file_path
        wb = load_workbook(file_path)
        if not wb.sheetnames:
            print("Error, the worksheet is not compatible, no sheet detected")
            sys.exit(1)

        sheet_name = "Rapport 1"
        sheet_name_input = input(
            "Enter the sheetname to read (default: " + sheet_name + "): "
        )
        sheet_name = sheet_name_input or sheet_name
        ws = wb[sheet_name]

        create_only = True
        print("Choose the action required (default 1) ")
        print("1: Create only, the already existing entry won't be updated")
        print(
            "2: Create and Update, create if it doesn't exist, "
            + "else update entry based on its pivots"
        )
        inp = input("Choose your option: ")

        if inp == "1":
            create_only = True
        elif inp == "2":
            create_only = False
        else:
            print("Using default option 1: Create only")

        # Create one object by row
        column_from_index = {}
        for my_tuple in ws["B4":"AH4"]:
            for cell in my_tuple:
                column_from_index[cell.column] = str(cell.value).strip()
        #        print(cell.value)

        # List here the fields to strip
        to_sprit = [
            "Nom Opération",
            "N° Opération GALION",
            "Commune",
            "Zone ABC",
            "Gestionnaire",
            "Gestionnaire (code)",
            "MOA (nom officiel)",
            "MOA (code SIRET)",
            "MOA Adresse 1",
            "MOA Code postal",
            "MOA Ville",
            "Type d'habitat",
            "Nom Opération",
            "Adresse Opération 1",
            "Opération code postal",
            "Nature logement",
            "Type Bénéficiaire",
            "Produit",
        ]

        to_sprit_input = map(
            lambda x: x.strip(),
            input(
                "List here the fields to strip (default: " + ",".join(to_sprit) + "): "
            ).split(","),
        )
        to_sprit = to_sprit_input or to_sprit

        my_objects = []
        my_parkings = []
        count_rows = 0
        count_inserted = 0
        for row in ws.iter_rows(
            min_row=5, max_row=ws.max_row, min_col=2, max_col=ws.max_column
        ):
            count_rows += 1
            my_row = {}
            my_row["count"] = count_rows + 4
            for cell in row:
                my_row[column_from_index[cell.column]] = (
                    str(cell.value).strip() if cell.column in to_sprit else cell.value
                )
            if len(my_row["MOA (code SIRET)"]) != 14:
                print(
                    f"le siret n'a pas de bon format, l'entrée est ignorée: {str_row(my_row)}"
                )
                continue
            if len(my_row["Gestionnaire (code)"]) != 5:
                print(
                    "le service instructeur n'est pas renseigné, "
                    + f"l'entrée est ignorée: {str_row(my_row)}"
                )
                continue
            if not my_row["Produit"] in dir(Financement):
                if my_row["Produit"][0:4] in ["PLUS", "PLAI"]:
                    my_row["Produit"] = my_row["Produit"][0:4]
                else:
                    print(
                        "le financement n'est pas supporté, "
                        + f"l'entrée est ignorée: {str_row(my_row)}"
                    )
                    continue
            my_objects.append(my_row)
            count_inserted += 1
            if (
                not my_row["Nb Garages aériens"] is None
                and int(my_row["Nb Garages aériens"]) > 0
            ):
                ga_row = my_row.copy()
                ga_row["Typologie Garage"] = "Garage Aérien"  # typologie
                ga_row["Nb Stationnement"] = my_row[
                    "Nb Garages aériens"
                ]  # nb_stationnements
                ga_row["Loyer"] = my_row["Loyer Garages aériens"]  # loyer
                my_parkings.append(ga_row)
            if (
                not my_row["Nb Garages enterrés"] is None
                and int(str(my_row["Nb Garages enterrés"])) > 0
            ):
                ge_row = my_row.copy()
                ge_row["Typologie Garage"] = "Garage enterré"  # typologie
                ge_row["Nb Stationnement"] = my_row[
                    "Nb Garages enterrés"
                ]  # nb_stationnements
                ge_row["Loyer"] = my_row["Loyer Garages enterrés"]  # loyer
                my_parkings.append(ge_row)
            if (
                not my_row["Nb Places Stationnement"] is None
                and int(str(my_row["Nb Places Stationnement"])) > 0
            ):
                ps_row = my_row.copy()
                ps_row["Typologie Garage"] = "Place de stationnement"  # typologie
                ps_row["Nb Stationnement"] = my_row[
                    "Nb Places Stationnement"
                ]  # nb_stationnements
                ps_row["Loyer"] = my_row["Loyer Places Stationnement"]  # loyer
                my_parkings.append(ps_row)
        print(f"{count_rows - count_inserted} / {count_rows} lignes ignorées")
        print(f"{len(my_objects)} lignes objet")
        print(f"{len(my_parkings)} lignes parking")

        Administration.map_and_create(my_objects, create_only)
        Bailleur.map_and_create(my_objects, create_only)
        Programme.map_and_create(my_objects, create_only)
        Lot.map_and_create(my_objects, create_only)
        TypeStationnement.map_and_create(my_parkings, create_only)


# Année Gestion Programmation
# Département
# N° Opération GALION
# Commune
# Zone 123
# Zone ABC
# Gestionnaire
# Gestionnaire (code)
# Gestionnaire (code SIREN)
# MOA (nom officiel)
# MOA (code SIRET)
# MOA Adresse 1
# MOA Code postal
# MOA Ville
# Type d'habitat
# Nom Opération
# Adresse Opération 1
# Opération code postal
# Année Achèvement Travaux
# Nature logement
# Type Bénéficiaire
# Produit
# Nb logts
# SU totale
# Nb Garages aériens
# Loyer Garages aériens
# Nb Garages enterrés
# Loyer Garages enterrés
# Nb Places Stationnement
# Loyer Places Stationnement
