import os
from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from instructeurs.models import Administration
from bailleurs.models import Bailleur

class Command(BaseCommand):
  def handle(self, **options):
    BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
    file_path = os.path.join(BASE_DIR, "documents", "v3.xlsx")
    val = input("Enter your value ("+file_path+"): ")
    if not(val):
      val = file_path
    wb = load_workbook(val)

    if not wb.sheetnames:
      print('Error, the worksheet is not compatible, no sheet detected')
      exit(1)
    sheet_name = 'Rapport 1'
    ws = wb[sheet_name]
    first_cell_of_table = 'B4'

    # Create one object by row
    column_from_index = {}
    for tuple in ws['B4':'AH4']:
      for cell in tuple:
        column_from_index[cell.column] = cell.value
#        print(cell.value)

    my_objects = []
    for row in ws.iter_rows(min_row=5,max_row=ws.max_row,min_col=2, max_col=ws.max_column): #(    'B{}:AH{}'.format(4,int(ws.max_row))):
      my_row = {}
      for cell in row:
        my_row[column_from_index[cell.column]] = cell.value
      if len(my_row['MOA (code SIRET)']) != 14:
        print(f"le siret n'a pas de bon format, l'entrée est ignorée: {my_row}")
        continue
      if not my_row['Gestionnaire (code)']:
        print(f"le service instructeur n'est pas renseigné, l'entrée est ignorée: {my_row}")
        continue
      my_objects.append(my_row)

    Administration.map_and_create(my_objects)
    Bailleur.map_and_create(my_objects)




    # admininstration_pivot= 'code'
    # admininstration_mapping= {
    #   'nom': 'Gestionnaire',
    #   'code': 'Gestionnaire (code)',
    # }

    # administrations = {}
    # for element in my_objects:
    #   element_pivot = element[admininstration_mapping[admininstration_pivot]]
    #   if not element_pivot in administrations:
    #     print(element_pivot)
    #     administrations[element_pivot] = {}
    #     for element_key in admininstration_mapping:
    #       administrations[element_pivot][element_key] = element[admininstration_mapping[element_key]]
    #     Administration.find_or_create_by_pivot(administrations[element_pivot])
    # administration2 = Administration.map_object(my_objects)
    # print(administration2)

# Année Gestion Programmation
# Département
# N° Opération GALION
# Commune
# Zone 123
# Zone ABC
# Type d'habitat
# Nom Opération
# Adresse Opération 1
# Opération code postal
# Année Achèvement Travaux
# Nature logement
# Type Bénéficiaire
# Produit
# Nb logts
# SU totale
# Nb Garages aériens
# Loyer Garages aériens
# Nb Garages enterrés
# Loyer Garages enterrés
# Nb Places Stationnement
# Loyer Places Stationnement


#    print(my_objects)

    # Close the workbook after reading
    wb.close()




# bailleur = {
#   'MOA (nom officiel)' : bailleur.nom,
#   'MOA (code SIRET)' : bailleur.siret,
#   'MOA Adresse 1' : bailleur.siege ?
#   'MOA Code postal' : bailleur.siege ?
#   'MOA Ville' : bailleur.siege ?
#   '' : bailleur.capital_social
#   '' : bailleur.dg_nom
#   '' : bailleur.dg_fonction
#   '' : bailleur.dg_date_deliberation
# }

# Programme = {
#   'Nom Opération': programme.nom,
#   ? : programme.code_postal,
#   'Commune': programme.ville,
#   'Adresse Opération 1': programme.adresse,
#   'Nb logts': programme.nb_logements,
#   'Zone 123': ?
#   'Zone ABC': ?
#   'Type Financement' : ? #beaucoup de mixte
#   'Produit' : ? # PLUS / PLAI/ PLS ... t plus encore
#   'Nature logement' ? # ordinaire, résidences, pension... etc
#   'Réf. Coordonnées XY': ?
#   'Type Bénéficiaire': ?, # Ménage Étidiants Jeunes...
#   'SU totale' : ?,
#   'Nb Jardins' : ?
#   'Loyer Jardins' : ?
#   ? : programme.type_operation,
#   ? : programme.anru,
#   ? : programme.nb_logement_non_conventionne,
#   ? : programme.nb_locaux_commerciaux,
#   ? : programme.nb_bureaux,
#   ? : programme.vendeur,
#   ? : programme.acquereur,
#   ? : programme.date_acte_notarie,
#   ? : programme.reference_notaire,
#   ? : programme.reference_publication_acte,
#   ? : programme.permis_construire,
#   'Année Gestion Programmation': programme.date_achevement_previsible, #différence année et date
#   ? : programme.date_achat,
#   'Année Achèvement Travaux': programme.date_achevement, #différence année et date -> célulle toujours vide
#   'Année Gestion Début': ?,
#   'Année Gestion Clôture': ?,
#   'N° Opération GALION': ?,
#   "Type d'habitat": ?, # individuel ou collectif
#   'Gestionnaire : ?': ? # lien avec l'instance instructeur ?
# }

# ReferenceCadastrale = {
#   'Réf. Parcelle cadastrale': ?
# }

# TypeStationnement = {
#   'Nb Garages aériens' : programme.Lot.TypeStationnement.topologie/nb_stationnements, # ici je n'ai pas de notion de lot
#   'Loyer Garages aériens' : programme.Lot.TypeStationnement.topologie/loyer, # ici je n'ai pas de notion de lot
#   'Nb Garages enterrés' : programme.Lot.TypeStationnement.topologie/nb_stationnements, # ici je n'ai pas de notion de lot
#   'Loyer Garages enterrés' : programme.Lot.TypeStationnement.topologie/loyer, # ici je n'ai pas de notion de lot
#   'Nb Places Stationnement' : programme.Lot.TypeStationnement.topologie/nb_stationnements, # ici je n'ai pas de notion de lot
#   'Loyer Places Stationnement' : programme.Lot.TypeStationnement.topologie/loyer, # ici je n'ai pas de notion de lot
# }