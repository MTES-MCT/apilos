from django.db.models import Q
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from bailleurs.models import Bailleur
from users.services import UserService


class BailleurListingProcessor:
    empty_lines_allowed = 5
    columns = {
        "first_name": ["prénom", "prenom"],
        "last_name": ["nom"],
        "email": [
            "email",
            "e-mail",
            "adresse email",
            "adresse e-mail",
            "courriel",
            "mail",
            "adresse mail",
        ],
        "bailleur": ["nom bailleur", "nom du bailleur", "bailleur"],
    }

    def __init__(self, filename):
        workbook: Workbook = load_workbook(filename=filename, data_only=True)
        self._worksheet: Worksheet = workbook[workbook.sheetnames[0]]

    def process(self) -> list[dict]:
        results = []
        # Create mapping:
        mapping = {}

        for column in self._worksheet.iter_cols(
            min_col=1, max_col=self._worksheet.max_column, min_row=1, max_row=1
        ):
            cell = column[0]
            if cell.value is None:
                continue
            for key, labels in self.columns.items():
                if cell.value.strip().lower() in labels:
                    mapping[key] = cell.column
                    break
            if len(mapping) == len(self.columns):
                break

        diff = self.columns.keys() - mapping.keys()
        keys = [f'"{self.columns[k][0]}"' for k in diff]
        if len(diff) > 0:
            if len(diff) == 1:
                raise Exception(
                    f"Lecture du fichier impossible: la colonne {keys[0]} est manquante"
                )
            raise Exception(
                f"Lecture du fichier impossible: les colonnes {', '.join(keys)} sont manquantes"
            )

        nb_empty_lines = 0

        for row in self._worksheet.iter_rows(
            min_col=1,
            max_col=self._worksheet.max_column,
            min_row=2,
            max_row=self._worksheet.max_row,
            values_only=True,
        ):
            data = {key: row[index - 1] for (key, index) in mapping.items()}

            # Check if row produces only empty data
            if all(v is None for v in data.values()):
                # If so check we haven't reached maximum number of allowed empty rows
                nb_empty_lines += 1
                if nb_empty_lines == self.empty_lines_allowed:
                    break
            else:
                bailleur = data.pop("bailleur")
                bailleur = str(bailleur).strip() if bailleur is not None else None
                data["bailleur"] = Bailleur.objects.filter(
                    Q(nom__iexact=bailleur) | Q(siret=bailleur)
                ).first()
                data["username"] = UserService.extract_username_from_email(
                    data.get("email", "")
                )
                results.append(data)

        return results
