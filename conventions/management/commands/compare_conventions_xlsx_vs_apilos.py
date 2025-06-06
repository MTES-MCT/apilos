import re
from collections.abc import Callable
from enum import Enum
from typing import Any

from django.contrib.postgres.search import TrigramSimilarity
from django.core.management.base import BaseCommand
from django.db.models import QuerySet
from django.db.models.functions import Lower
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from conventions.models.convention import Convention

SUCCESS_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
# WARNING_FILL = PatternFill(
#     start_color="FFA500", end_color="FFA500", fill_type="solid"
# )
ERROR_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Init fields from workbook
# Only those fields will be reported in output sheet
MAPPING_HEADERS = {
    "numero": "N°convention",
    "annee_convention": "Année convention",
    "adresse": "Adresse",
    "ville": "Communes",
    "code_insee_commune": "INSEE",
    "bailleur": "Bailleurs",
    "nom": "Nom de l’opération",
    "financement": "financement",
    "nb_logements": "Nb de logts",
    "avenant": "Avenant oui/non",
}

RESULT_HEADERS = {
    "numero": "Numero dans APiLos",
    "annee_convention": "Année de la convention dans APiLos",
    "programme__adresse": "Adresse du programme dans APiLos",
    "programme__ville": "Ville du programme dans APiLos",
    "programme__code_insee_commune": "Code commune dans APiLos",
    "programme__bailleur": "Bailleur du programme dans APiLos",
    "programme__nom": "Nom du programme dans APiLos",
    "lot__financement": "Financement dans APiLos",
    "lot__nb_logements": "Nb de lgts dans APiLos",
    "statut": "Statut dans APiLos",
}

RESULT_COMPUTED_HEADERS = {
    "avenant": "Avenant dans APiLos",
    "methode": "Méthode de recherche de la convention",
}

# Metadata keys
NB_CONV_IN_FILE = "Nombre de conventions dans le fichier"
NB_CONV_WITH_NUMBER = "Nombre de conventions dans le fichier avec numéro"
NB_CONV_FOUND = "Nombre de conventions trouvées"
NB_CONV_IN_DB = "Nombre de conventions dans la base de données"


class FindRule:
    def __init__(self, name: str, description: str, method: Callable):
        self.name = name
        self.description = description
        self.method = method


def find_by_numero(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    if numero := _number_for_search(convention[MAPPING_HEADERS["numero"]]):

        convention_from_db = convention_qs.filter(numero_pour_recherche=numero)
        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le numéro {numero}"
                )
            )
    return None

    # NUMERO_WITHOUT_SIXTH_NUMBER = "numero sans le 6ème chiffre"


def find_by_figure_13_first_and_3_last(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    numero = convention[MAPPING_HEADERS["numero"]]

    if numero is None:
        return None
    # garder que les chiffres
    numero_array = re.findall(r"\d+", numero)
    numero_array = "".join(numero_array)
    if len(numero_array) < 16:
        return None
    numero_startwith = numero_array[:13]
    numero_endwith = numero_array[-3:]

    convention_from_db = convention_qs.filter(
        numero_pour_recherche__startswith=numero_startwith,
        numero_pour_recherche__endswith=numero_endwith,
    )
    if len(convention_from_db) == 1:
        cmd.stdout.write(
            cmd.style.SUCCESS(
                f"Convention trouvée pour le numéro {numero}, "
                f"recherche sur le début du numero {numero_startwith}, "
                f"recherche sur la fin du numero {numero_endwith}"
            )
        )
        return convention_from_db[0]
    elif len(convention_from_db) > 1:
        cmd.stdout.write(
            cmd.style.WARNING(
                f"Plusieurs conventions trouvées pour le numéro {numero}, "
                f"recherche sur le début du numero {numero_startwith}, "
                f"recherche sur la fin du numero {numero_endwith}"
            )
        )
    return None


def find_by_numero_without_sixth_number(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    numero = convention[MAPPING_HEADERS["numero"]]
    if numero is None:
        return None
    numero_array = numero.split(" ")
    if len(numero_array) < 7:
        return None
    if int(numero_array[5]) >= 10:
        return None

    numero_startwith = "".join(numero_array[:5])
    numero_endwith = numero_array[-1]

    convention_from_db = convention_qs.filter(
        numero_pour_recherche__startswith=numero_startwith,
        numero_pour_recherche__endswith=numero_endwith,
    )
    if len(convention_from_db) == 1:
        cmd.stdout.write(
            cmd.style.SUCCESS(
                f"Convention trouvée pour le numéro {numero}, "
                f"recherche sur le début du numero {numero_startwith}, "
                f"recherche sur la fin du numero {numero_endwith}"
            )
        )
        return convention_from_db[0]
    elif len(convention_from_db) > 1:
        cmd.stdout.write(
            cmd.style.WARNING(
                f"Plusieurs conventions trouvées pour le numéro {numero}, "
                f"recherche sur le début du numero {numero_startwith}, "
                f"recherche sur la fin du numero {numero_endwith}"
            )
        )
    return None


def find_by_numero_endwith_3_figures_code(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]
    numero = convention[MAPPING_HEADERS["numero"]]
    # get the last number
    if not numero:
        return None
    numero_array = numero.replace(" ", "")
    numero_endwith = numero_array[-3:]

    if numero_endwith := _number_for_search(numero_endwith):
        convention_from_db = convention_qs.filter(
            numero_pour_recherche__endswith=numero_endwith,
            programme__code_insee_commune=code_insee_commune,
        )
        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le numéro {numero}"
                    f", recherche sur la fin du numero {numero_endwith}"
                    f" et code commune {code_insee_commune}"
                )
            )
    return None


def find_by_numero_endwith_3_figures_commune(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    commune = convention["commune_new"]
    numero = convention[MAPPING_HEADERS["numero"]]
    # get the last number
    if not numero:
        return None
    numero_array = numero.replace(" ", "")
    numero_endwith = numero_array[-3:]

    if numero_endwith := _number_for_search(numero_endwith):
        convention_from_db = convention_qs.filter(
            numero_pour_recherche__endswith=numero_endwith,
            programme__ville__iexact=commune,
        )
        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le numéro {numero}"
                    f", recherche sur la fin du numero {numero_endwith}"
                    f" et commune {commune}"
                )
            )
    return None


# NUMERO_ENDWITH__CODE = "numero fini par et code commune"
def find_by_numero_endwith_code(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]
    numero = convention[MAPPING_HEADERS["numero"]]
    # get the last number
    if not numero:
        return None
    numero_array = numero.split(" ")
    numero_endwith = numero_array[-1]

    if numero_endwith := _number_for_search(numero_endwith):
        convention_from_db = convention_qs.filter(
            numero_pour_recherche__endswith=numero_endwith,
            programme__code_insee_commune=code_insee_commune,
        )
        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le numéro {numero}"
                    f", recherche sur la fin du numero {numero_endwith}"
                    f" et code commune {code_insee_commune}"
                )
            )
    return None


# NOM_EXACT__FIN__CODE = (
#     "nom, financement et code commune : match exact"
# )
def find_by_nom_exact_fin_code(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    nom = convention[MAPPING_HEADERS["nom"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]

    if nom:
        convention_from_db = convention_qs.filter(
            programme__nom__iexact=nom,
            lots__financement=financement,
            programme__code_insee_commune=code_insee_commune,
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le nom {nom}, financement"
                    f" {financement} et code commune {code_insee_commune}"
                )
            )
    return None


# NOM_TRIGRAM_08__FIN__CODE = (
#     "nom (similairité > 0.8), financement et code commune"
# )


def find_by_nom_trigram_fin_code(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    nom = convention[MAPPING_HEADERS["nom"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]

    if nom:
        convention_from_db = (
            convention_qs.annotate(
                similarity=TrigramSimilarity(Lower("programme__nom"), nom.lower())
            )
            .filter(
                similarity__gt=0.8,
                lots__financement=financement,
                programme__code_insee_commune=code_insee_commune,
            )
            .order_by("-similarity")
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le nom {nom}"
                    f" (trigram > 0.8), financement {financement} et"
                    f" code commune {code_insee_commune}"
                )
            )
    return None


# ADRESSE_EXACT__FIN__CODE = (
#     "adresse, financement et code commune : match exact"
# )
def find_by_adresse_exact_fin_code(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    adresse = convention[MAPPING_HEADERS["adresse"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]

    if adresse:
        convention_from_db = convention_qs.filter(
            programme__adresse__iexact=adresse,
            lots__financement=financement,
            programme__code_insee_commune=code_insee_commune,
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le adresse {adresse},"
                    f" financement {financement} et"
                    f" code commune {code_insee_commune}"
                )
            )
    return None


# ADRESSE_TRIGRAM_08__FIN__CODE = (
#     "adresse (similairité > 0.8), financement et code commune : trigramme 0.8"
# )
def find_by_adresse_trigram_fin_code(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    adresse = convention[MAPPING_HEADERS["adresse"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]
    if adresse:
        convention_from_db = (
            convention_qs.annotate(
                similarity=TrigramSimilarity(
                    Lower("programme__adresse"), adresse.lower()
                )
            )
            .filter(
                similarity__gt=0.8,
                lots__financement=financement,
                programme__code_insee_commune=code_insee_commune,
            )
            .order_by("-similarity")
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]

        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le adresse {adresse}"
                    f" (trigram > 0.8), financement {financement} et"
                    f" code commune {code_insee_commune}"
                )
            )
    return None


# ADRESSE_WITH_NOM_EXACT__CODE__FIN__NB_LGTS = (
#     "adresse (fichier) avec le nom (APILOS), financement, nb logements"
# )
def find_by_adresse_with_nom_exact_code_fin_nb_lgts(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    adresse = convention[MAPPING_HEADERS["adresse"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]

    if adresse:
        convention_from_db = convention_qs.filter(
            programme__nom__iexact=adresse,
            lots__financement=financement,
            programme__code_insee_commune=code_insee_commune,
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]
        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le adresse {adresse},"
                    f" financement {financement} et"
                    f" code commune {code_insee_commune}"
                )
            )
    return None


# ADRESSE_WITH_NOM_TRIGRAM_08__CODE__FIN__NB_LGTS = (
#     "adresse (fichier) avec le nom (APILOS) (similairité > 0.8), financement,
#     nb de logements"
# )
def find_by_adresse_with_nom_trigram_code_fin_nb_lgts(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    adresse = convention[MAPPING_HEADERS["adresse"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]
    if adresse:
        convention_from_db = (
            convention_qs.annotate(
                similarity=TrigramSimilarity(Lower("programme__nom"), adresse.lower())
            )
            .filter(
                similarity__gt=0.6,
                lots__financement=financement,
                programme__code_insee_commune=code_insee_commune,
            )
            .order_by("-similarity")
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]

        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le adresse {adresse}"
                    f" (trigram > 0.6), financement {financement} et"
                    f" code commune {code_insee_commune}"
                )
            )
    return None


# ADRESSE_WITH_NOM_TRIGRAM_08__CODE__NB_LGTS = (
#     "adresse (fichier) avec le nom (APILOS) (similairité > 0.8), financement,
#     nb de logements"
# )
def find_by_adresse_with_nom_trigram_code_nb_lgts(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    adresse = convention[MAPPING_HEADERS["adresse"]]
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]
    if adresse:
        convention_from_db = (
            convention_qs.annotate(
                similarity=TrigramSimilarity(Lower("programme__nom"), adresse.lower())
            )
            .filter(
                similarity__gt=0.6,
                programme__code_insee_commune=code_insee_commune,
            )
            .order_by("-similarity")
        )

        if len(convention_from_db) == 1:
            return convention_from_db[0]

        elif len(convention_from_db) > 1:
            cmd.stdout.write(
                cmd.style.WARNING(
                    f"Plusieurs conventions trouvées pour le adresse {adresse}"
                    f" (trigram > 0.6), financement {financement} et"
                    f" code commune {code_insee_commune}"
                )
            )
    return None


# FIN_CODE_NB_LGTS = "financement, code commune, nb de logements"
def find_by_financement_code_nb_lgts(
    cmd: BaseCommand, convention: dict, convention_qs: QuerySet
) -> Convention | None:
    financement = convention[MAPPING_HEADERS["financement"]]
    code_insee_commune = convention[MAPPING_HEADERS["code_insee_commune"]]
    nb_logements = convention[MAPPING_HEADERS["nb_logements"]]
    convention_from_db = (
        convention_qs.filter(
            lots__financement=financement,
            programme__code_insee_commune=code_insee_commune,
        )
        .filter(lots__nb_logements=nb_logements)
        .distinct()
    )

    if len(convention_from_db) == 1:
        return convention_from_db[0]
    elif len(convention_from_db) > 1:
        cmd.stdout.write(
            cmd.style.WARNING(
                f"Plusieurs conventions trouvées pour le financement {financement},"
                f" code commune {code_insee_commune} et nb_logements {nb_logements}"
            )
        )
    return None


class FindRuleBy(Enum):

    # 1098
    NUMERO = FindRule("NUMERO", "Trouvé à partir du numero", find_by_numero)

    # 2172
    FIGURE_13_FIRST_AND_3_LAST = FindRule(
        "FIGURE_13_FIRST_AND_3_LAST",
        "Trouvé à partir des 13 premiers et des 3 derniers chiffres",
        find_by_figure_13_first_and_3_last,
    )

    # 0 -> FIXME : Useless ?
    NUMERO_WITHOUT_SIXTH_NUMBER = FindRule(
        "NUMERO_WITHOUT_SIXTH_NUMBER",
        "Trouvé à partir du numero sans le 6ème chiffre",
        find_by_numero_without_sixth_number,
    )

    # Nom de la commune (commune_new dans notre tableur) + les 3 derniers chiffres de la convention
    # Code INSEE (insee_new) + année + bailleur
    # Nom de la commune (commune_new dans notre tableur) + année + bailleur

    # 153
    NUMERO_ENDWITH_3_FIGURES_CODE = FindRule(
        "NUMERO_ENDWITH_3_FIGURES_CODE",
        "Trouvé à partir du numero fini par 3 chiffres et code commune",
        find_by_numero_endwith_3_figures_code,
    )

    # Nom de la commune (commune_new dans notre tableur) + les 3 derniers chiffres de la convention
    NUMERO_ENDWITH_3_FIGURES_COMMUNE = FindRule(
        "NUMERO_ENDWITH_3_FIGURES_CODE",
        "Trouvé à partir du numero fini par 3 chiffres et nom de la commune",
        find_by_numero_endwith_3_figures_commune,
    )

    # 0 -> FIXME : Useless ?
    # NUMERO_ENDWITH__CODE = FindRule(
    #     "NUMERO_ENDWITH__CODE",
    #     "Trouvé à partir du numero fini par et code commune",
    #     find_by_numero_endwith_code,
    # )

    # 8 -> FIXME : Faux positif ?
    NOM_EXACT__FIN__CODE = FindRule(
        "NOM_EXACT__FIN__CODE",
        "Trouvé à partir du nom (exact), financement et code commune",
        find_by_nom_exact_fin_code,
    )

    # 0 -> FIXME : Useless ?
    NOM_TRIGRAM_08__FIN__CODE = FindRule(
        "NOM_TRIGRAM_08__FIN__CODE",
        "Trouvé à partir du nom (similairité > 0.8), financement et code commune",
        find_by_nom_trigram_fin_code,
    )

    # 8 -> FIXME : Faux positif ?
    ADRESSE_EXACT__FIN__CODE = FindRule(
        "ADRESSE_EXACT__FIN__CODE",
        "Trouvé à partir de l'adresse (exact), financement et code commune",
        find_by_adresse_exact_fin_code,
    )

    # 12 -> FIXME : Faux positif ?
    ADRESSE_TRIGRAM_08__FIN__CODE = FindRule(
        "ADRESSE_TRIGRAM_08__FIN__CODE",
        "Trouvé à partir de l'adresse (similairité > 0.8), financement et code commune",
        find_by_adresse_trigram_fin_code,
    )

    # 0 -> FIXME : Useless ?
    ADRESSE_WITH_NOM_EXACT__CODE__FIN__NB_LGTS = FindRule(
        "ADRESSE_WITH_NOM_EXACT__CODE__FIN__NB_LGTS",
        "Trouvé à partir de l'adresse (fichier) comparée avec le nom (APILOS),"
        " financement, nb logements",
        find_by_adresse_with_nom_exact_code_fin_nb_lgts,
    )

    # 8 -> FIXME : Faux positif ?
    ADRESSE_WITH_NOM_TRIGRAM_06__CODE__FIN__NB_LGTS = FindRule(
        "ADRESSE_WITH_NOM_TRIGRAM_06__CODE__FIN__NB_LGTS",
        "Trouvé à partir de l'adresse (fichier) comparée avec le nom (APILOS)"
        " (similairité > 0.6), financement, nb de logements",
        find_by_adresse_with_nom_trigram_code_fin_nb_lgts,
    )

    # 18 -> FIXME : Faux positif ?
    ADRESSE_WITH_NOM_TRIGRAM_06__CODE__NB_LGTS = FindRule(
        "ADRESSE_WITH_NOM_TRIGRAM_06__CODE__NB_LGTS",
        "Trouvé à partir de l'adresse (fichier) comparée avec le nom (APILOS)"
        " (similairité > 0.6), nb de logements",
        find_by_adresse_with_nom_trigram_code_nb_lgts,
    )

    # 39 -> FIXME : Faux positif ?
    FIN_CODE_NB_LGTS = FindRule(
        "FIN_CODE_NB_LGTS",
        "Trouvé à partir du financement, code commune, nb de logements",
        find_by_financement_code_nb_lgts,
    )

    @classmethod
    def get_find_rules(cls) -> list[FindRule]:
        return [rule.value for rule in cls.__members__.values()]


def _init_metadata() -> dict[str, int]:
    metadata = {
        NB_CONV_IN_FILE: 0,
        NB_CONV_WITH_NUMBER: 0,
        NB_CONV_FOUND: 0,
        NB_CONV_IN_DB: 0,
    }
    for value in FindRuleBy.__members__.values():
        metadata[value.value.description] = 0
    return metadata


def _number_for_search(number: Any) -> str:
    if not number or not isinstance(number, str):
        return ""
    return (
        number.replace(" ", "")
        .replace("-", "")
        .replace(".", "")
        .replace(":", "")
        .replace("_", "")
    )


def _set_header(output_wb_sheet: Worksheet, apilos_wb_sheet: Worksheet) -> None:

    column_nb = 1
    for header in MAPPING_HEADERS.values():
        output_wb_sheet.cell(row=1, column=column_nb, value=header)
        column_nb += 1

    for header in RESULT_HEADERS.values():
        output_wb_sheet.cell(row=1, column=column_nb, value=header)
        column_nb += 1

    for header in RESULT_COMPUTED_HEADERS.values():
        output_wb_sheet.cell(row=1, column=column_nb, value=header)
        column_nb += 1

    column_nb = 1
    for header in RESULT_HEADERS.values():
        apilos_wb_sheet.cell(row=1, column=column_nb, value=header)
        column_nb += 1


def row_to_dict(
    row: tuple[Cell], column_from_index: dict, output_wb_sheet: Worksheet
) -> tuple[int, int, dict]:
    convention = {}
    column_nb = 1
    row_nb = 0
    for cell in row:
        if cell.column > len(column_from_index):
            break
        row_nb = cell.row
        convention[column_from_index[cell.column]] = cell.value

    for field in MAPPING_HEADERS.values():
        output_wb_sheet.cell(row=cell.row, column=column_nb, value=convention[field])
        column_nb += 1
    return row_nb, column_nb, convention


def write_metadata(metadata_wb_sheet, metadata):
    row = 1
    for kpi, value in metadata.items():
        metadata_wb_sheet.cell(row=row, column=1, value=kpi)
        metadata_wb_sheet.cell(row=row, column=2, value=value)
        row += 1


def write_not_found_conventions(apilos_wb_sheet, convention_qs, found_convention_ids):

    row_nb = 2
    for convention in convention_qs.all():
        # doute sur le exclude
        if convention.id in found_convention_ids:
            continue

        last_convention_version = (
            convention.avenants.all().order_by("-cree_le").first()
        ) or convention

        column_nb = 1
        for field in RESULT_HEADERS.keys():
            if field == "numero":
                conv = convention
            else:
                conv = last_convention_version

            compose = field.split("__")
            if len(compose) == 2:
                conv = getattr(conv, compose[0])
                field = compose[1]

            value = getattr(conv, field, "") if conv else ""
            value = str(value) if field == "bailleur" else value
            apilos_wb_sheet.cell(
                row=row_nb,
                column=column_nb,
                value=value,
            )
            column_nb += 1
        row_nb += 1


def write_matching_conventions(
    cmd: BaseCommand,
    input_wb_sheet,
    output_wb_sheet,
    convention_qs,
    column_from_index,
    metadata,
):
    found_convention_ids = []

    def _find_convention_by(
        cmd, convention: dict, convention_qs: QuerySet
    ) -> tuple[Convention | None, FindRule | None]:

        for find_rule in FindRuleBy.get_find_rules():
            if convention_from_db := find_rule.method(cmd, convention, convention_qs):
                return convention_from_db, find_rule

        return None, None

    # Read input sheet
    for row in input_wb_sheet.iter_rows(min_row=2):
        row_nb, column_nb, convention = row_to_dict(
            row, column_from_index, output_wb_sheet
        )
        metadata[NB_CONV_IN_FILE] += 1
        convention_from_db, find_by = _find_convention_by(
            cmd, convention, convention_qs
        )
        if (
            convention[MAPPING_HEADERS["numero"]]
            and convention[MAPPING_HEADERS["numero"]].strip()
        ):
            metadata[NB_CONV_WITH_NUMBER] += 1

        if find_by:
            metadata[find_by.description] += 1

        if convention_from_db and find_by:
            metadata[NB_CONV_FOUND] += 1
            found_convention_ids.append(convention_from_db.id)
            last_convention_version = (
                convention_from_db.avenants.all().order_by("-cree_le").first()
            ) or convention_from_db

            for field in RESULT_HEADERS.keys():
                if field == "numero":
                    conv = convention_from_db
                else:
                    conv = last_convention_version

                compose = field.split("__")
                if len(compose) == 2:
                    conv = getattr(conv, compose[0])
                    field = compose[1]

                value = getattr(conv, field, "") if conv else ""
                value = str(value) if field == "bailleur" else value
                cell = output_wb_sheet.cell(
                    row=row_nb,
                    column=column_nb,
                    value=value,
                )
                column_nb += 1
                if field in MAPPING_HEADERS and MAPPING_HEADERS[field] in convention:
                    if convention[MAPPING_HEADERS[field]] == value:
                        cell.fill = SUCCESS_FILL
                    else:
                        cell.fill = ERROR_FILL

            has_avenant = (
                "Oui" if convention_from_db != last_convention_version else "Non"
            )
            cell = output_wb_sheet.cell(
                row=row_nb,
                column=column_nb,
                value=has_avenant,
            )
            if has_avenant != convention[MAPPING_HEADERS["avenant"]]:
                cell.fill = ERROR_FILL
            else:
                cell.fill = SUCCESS_FILL

            column_nb += 1
            output_wb_sheet.cell(
                row=row_nb,
                column=column_nb,
                value=find_by.description if find_by else "",
            )
    found_convention_ids = list(set(found_convention_ids))

    return found_convention_ids


class Command(BaseCommand):
    help = """
Read the xlsx file of exhaustive list of conventions and compare with the
conventions in the database
"""

    def _should_continue(self) -> None:
        self.stdout.write(
            self.style.NOTICE(
                f"""
Lecture du fichier : {self.conv_file}
Pour le département : {self.department}
Feuille de calcule à interpréter : {self.input_sheetname}
Feuille de calcule contenant le résultat : {self.output_sheetname}
(Si une feuille de calcule `{self.output_sheetname}` existe, elle sera ré-initialisée)
Feuille de calcule contenant les indicateurs globales : {self.metadata_sheetname}
(Si une feuille de calcule `{self.metadata_sheetname}` existe, elle sera ré-initialisée)
                """
            )
        )

        while True:
            result = input("continuer ? (y/n)")
            if result.lower() == "n":
                self.stdout.write(
                    self.style.NOTICE(
                        "Vous avez choisi de stopper l'execution de cette commande"
                    )
                )
                exit(1)
            if result.lower() == "y":
                break

    def _find_convention_by(
        self, convention: dict, convention_qs: QuerySet
    ) -> tuple[Convention | None, FindRule | None]:

        for find_rule in FindRuleBy.get_find_rules():
            if convention_from_db := find_rule.method(self, convention, convention_qs):
                return convention_from_db, find_rule

        return None, None

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Convention exhauustive list",
        )

    def _get_wb_sheet(
        self, conv_workbook: Workbook
    ) -> tuple[Worksheet, Worksheet, Worksheet, Worksheet]:

        for sheetname in [
            self.output_sheetname,
            self.apilos_sheetname,
            self.metadata_sheetname,
        ]:
            if sheetname in conv_workbook:
                del conv_workbook[sheetname]
            conv_workbook.create_sheet(sheetname)

        # Get output and metadata sheets
        output_wb_sheet = conv_workbook[self.output_sheetname]
        apilos_wb_sheet = conv_workbook[self.apilos_sheetname]
        metadata_wb_sheet = conv_workbook[self.metadata_sheetname]

        # Get input sheet
        try:
            input_wb_sheet = conv_workbook[self.input_sheetname]
        except KeyError:
            self.stdout.write(
                self.style.ERROR(
                    f"la feuille de calcule {self.input_sheetname} n'existe pas"
                    f" dans le fichier {self.conv_file}"
                )
            )
            exit(1)
        return input_wb_sheet, output_wb_sheet, apilos_wb_sheet, metadata_wb_sheet

    def handle(self, *args, **options):
        self.conv_file = options["file"]
        self.department = 79
        self.input_sheetname = "Inventaire convention"
        self.output_sheetname = "Resultats"
        self.metadata_sheetname = "Metadonnées"
        self.apilos_sheetname = "Conventions APiLos non trouvées dans le fichier"

        # Display what the command will do
        self._should_continue()

        convention_qs = Convention.objects.filter(
            programme__code_insee_departement=self.department, parent_id__isnull=True
        ).prefetch_related("lots", "programme", "programme__bailleur")
        metadata: dict[str, int] = _init_metadata()
        metadata[NB_CONV_IN_DB] = convention_qs.count()

        # Get input and output workbook sheets
        try:
            conv_workbook = load_workbook(filename=self.conv_file, data_only=True)
        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f"le fichier {self.conv_file} n'existe pas")
            )
            exit(1)

        # Init output and metadata sheets
        input_wb_sheet, output_wb_sheet, apilos_wb_sheet, metadata_wb_sheet = (
            self._get_wb_sheet(conv_workbook)
        )

        column_from_index = {}
        # read the first row (header)
        for first_row in input_wb_sheet.iter_rows(min_row=1, max_row=1):
            for cell in first_row:
                if cell.value is None:
                    break
                column_from_index[cell.column] = str(cell.value).strip()

        self.stdout.write(f" Entêtes : {column_from_index.values()}")

        _set_header(output_wb_sheet, apilos_wb_sheet)

        found_convention_ids = write_matching_conventions(
            self,
            input_wb_sheet,
            output_wb_sheet,
            convention_qs,
            column_from_index,
            metadata,
        )

        write_not_found_conventions(
            apilos_wb_sheet, convention_qs, found_convention_ids
        )

        write_metadata(metadata_wb_sheet, metadata)

        conv_workbook.save(filename=self.conv_file)
