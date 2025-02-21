import datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

from django.core.exceptions import FieldDoesNotExist
from openpyxl import load_workbook
from zipfile import BadZipFile

from core.storage import client
from programmes.models import TypologieLogement

from . import utils


def _save_uploaded_file(my_file, convention, file_name):
    my_file.seek(0)
    client.put_object(
        my_file=my_file.read(),
        target=f"conventions/{convention.uuid}/uploads/{file_name}",
    )


def handle_uploaded_xlsx(
    upform,
    my_file,
    my_class,
    convention,
    file_name,
    class_field_mapping="import_mapping",
    class_field_needed_mapping="needed_in_mapping",
    import_order=False,
):
    try:
        my_file.seek(0)
        my_wb = load_workbook(filename=BytesIO(my_file.read()), data_only=True)
    except BadZipFile:
        upform.add_error(
            "file",
            "Le fichier importé ne semble pas être du bon format, 'xlsx' est le format attendu",
        )
        return {"success": utils.ReturnStatus.ERROR}
    except KeyError:
        upform.add_error(
            "file",
            "Il y a probablement des fichiers manquants dans l'archive",
        )
        return {"success": utils.ReturnStatus.ERROR}
    except ValueError:
        upform.add_error(
            "file",
            "Certains fichiers de l'archive ont probablement un contenu invalide",
        )
        return {"success": utils.ReturnStatus.ERROR}

    try:
        my_ws = my_wb[my_class.sheet_name]
    except KeyError:
        upform.add_error(
            "file",
            f"Le fichier importé doit avoir une feuille nommée '{my_class.sheet_name}'",
        )
        return {"success": utils.ReturnStatus.ERROR}

    min_row = 3
    if "Data" in my_wb:
        if my_wb["Data"]["E2"].value:
            min_row = int(my_wb["Data"]["E2"].value)

    if convention is not None:
        _save_uploaded_file(my_file, convention, file_name)

    column_from_index, import_warnings = _check_not_useful_columns(
        my_ws, my_class, class_field_mapping
    )

    if _has_error_column_header_exist(
        upform, my_class, class_field_mapping, column_from_index
    ):
        return {"success": utils.ReturnStatus.ERROR}

    # transform each line into object
    my_objects, import_warnings = _get_object_from_worksheet(
        my_ws,
        column_from_index,
        my_class,
        import_warnings,
        min_row,
        class_field_mapping=class_field_mapping,
        class_field_needed_mapping=class_field_needed_mapping,
        import_order=import_order,
    )

    return {
        "success": (
            utils.ReturnStatus.SUCCESS
            if len(import_warnings) == 0
            else utils.ReturnStatus.WARNING
        ),
        "objects": my_objects,
        "import_warnings": import_warnings,
    }


def _check_not_useful_columns(my_ws, my_class, class_field_mapping):
    import_warnings = []
    column_from_index = {}
    for col in my_ws.iter_cols(
        min_col=1, max_col=my_ws.max_column, min_row=1, max_row=1
    ):
        for cell in col:
            if cell.value is None:
                continue
            key = all_words_in_key_of_dict(
                cell.value, getattr(my_class, class_field_mapping)
            )
            if key is None:
                import_warnings.append(
                    Exception(
                        f"La colonne nommée '{cell.value}' est inconnue, "
                        + "elle sera ignorée. Le contenu des colonnes "
                        + "attendus est dans la liste : "
                        + f"{', '.join(getattr(my_class, class_field_mapping).keys())}"
                    )
                )
                continue
            column_from_index[cell.column] = key
    return column_from_index, import_warnings


def _has_error_column_header_exist(
    upform, my_class, class_field_mapping, column_from_index
):
    error_column = False
    for key in getattr(my_class, class_field_mapping):
        if not all_words_in_key_of_dict(key, list(column_from_index.values())):
            upform.add_error(
                "file",
                (
                    "Le fichier importé doit avoir une entête de colonne avec "
                    + f"le contenu '{key}'"
                ),
            )
            error_column = True
    return error_column


def _get_object_from_worksheet(
    my_ws,
    column_from_index,
    my_class,
    import_warnings,
    min_row=3,
    *,
    class_field_mapping,
    class_field_needed_mapping,
    import_order: bool,
):
    my_objects = []

    for index, row in enumerate(
        my_ws.iter_rows(
            min_row=min_row, max_row=my_ws.max_row, min_col=1, max_col=my_ws.max_column
        )
    ):
        my_row, empty_line, new_warnings = _extract_row(
            row, column_from_index, my_class, class_field_mapping=class_field_mapping
        )

        if import_order:
            my_row["import_order"] = index

        if hasattr(my_class, class_field_needed_mapping):
            if not empty_line and getattr(my_class, class_field_needed_mapping):
                for needed_field in getattr(my_class, class_field_needed_mapping):
                    if needed_field not in my_row:
                        empty_line = True
                        new_warnings.append(
                            Exception(
                                f"La ligne {row[0].row} a été ignorée car la"
                                + f" valeur '{needed_field}'"
                                + " n'est pas renseignée"
                            )
                        )
                        break

        import_warnings = [*import_warnings, *new_warnings]

        # Ignore if the line is empty
        if not empty_line:
            my_objects.append(my_row)
    return my_objects, import_warnings


def _extract_row(row, column_from_index, cls, *, class_field_mapping):
    # pylint: disable=R0912
    new_warnings = []
    my_row = {}
    empty_line = True
    import_mapping = getattr(cls, class_field_mapping)
    for cell in row:
        # Ignore unknown column
        if cell.column not in column_from_index or cell.value is None:
            continue

        # Check the empty lines to don't fill it
        empty_line = False
        value = None
        try:
            model_field = cls._meta.get_field(
                import_mapping[column_from_index[cell.column]]
            )
        except FieldDoesNotExist:
            model_field = import_mapping[column_from_index[cell.column]]

        if isinstance(model_field, str):
            key = model_field
            value = cell.value
        else:
            key = model_field.name

            # Date case
            if model_field.get_internal_type() == "DateField":
                if isinstance(cell.value, datetime.datetime):
                    value = utils.format_date_for_form(cell.value)
                else:
                    new_warnings.append(
                        Exception(
                            f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                            + f"de la colonne {column_from_index[cell.column]} "
                            + "doit être une date"
                        )
                    )

            # TextChoices case
            elif (
                model_field.get_internal_type() == "CharField"
                and model_field.choices is not None
            ):
                if cell.value is not None:
                    tmp_value = cell.value
                    if model_field.choices == TypologieLogement.choices:
                        tmp_value = TypologieLogement.map_string(tmp_value)
                    value = next(
                        (x[0] for x in model_field.choices if x[1] == tmp_value), None
                    )
                    if (
                        value is None
                    ):  # value is not Null but not in the choices neither
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit faire partie des valeurs : "
                                + f"{', '.join(map(lambda x : x[1], model_field.choices))}"
                            )
                        )

            # Float case
            elif model_field.get_internal_type() == "FloatField":
                if cell.value is not None:
                    if isinstance(cell.value, (float, int)):
                        value = float(cell.value)
                    else:
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit être une valeur numérique"
                            )
                        )

            # Decimal case
            elif model_field.get_internal_type() == "DecimalField":
                value, new_warnings = _get_value_from_decimal_field(
                    cell, model_field, column_from_index, new_warnings
                )

            # Integer case
            elif model_field.get_internal_type() == "IntegerField":
                if cell.value is not None:
                    if isinstance(cell.value, (float, int)):
                        value = int(cell.value)
                    else:
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit être une valeur numérique"
                            )
                        )

            # String case
            elif model_field.get_internal_type() == "CharField":
                if cell.value is not None:
                    if isinstance(cell.value, (float, int, str)):
                        value = cell.value
                    else:
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit être une valeur alphanumeric"
                            )
                        )
        my_row[key] = value

    return my_row, empty_line, new_warnings


def _float_to_decimal_rounded_half_up(number: float, decimal_places: int) -> Decimal:
    if decimal_places > 0:
        decimal_target = "1." + ("0" * decimal_places)
    else:
        decimal_target = "1"
    return Decimal(str(number)).quantize(
        Decimal(decimal_target), rounding=ROUND_HALF_UP
    )


def _get_value_from_decimal_field(cell, model_field, column_from_index, new_warnings):
    value = None
    if cell.value is not None:
        if isinstance(cell.value, str):
            try:
                local_format = "{:." + str(model_field.decimal_places) + "f}"
                value = Decimal(
                    local_format.format(_extract_float_from_string(cell.value))
                )
            except ValueError:
                new_warnings.append(
                    Exception(
                        f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                        + f"de la colonne {column_from_index[cell.column]} "
                        + "doit être une valeur numérique"
                    )
                )
        elif isinstance(cell.value, (float, int)):
            value = _float_to_decimal_rounded_half_up(
                cell.value, model_field.decimal_places
            )
        else:
            new_warnings.append(
                Exception(
                    f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                    + f"de la colonne {column_from_index[cell.column]} "
                    + "doit être une valeur numérique"
                )
            )
    return value, new_warnings


def _extract_float_from_string(my_string: str):
    my_string = my_string.strip()
    my_string = my_string.replace(",", ".")
    i = 0
    for char in my_string:
        if char in ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "."]:
            i += 1
        else:
            break
    return float(my_string[:i])


def all_words_in_key_of_dict(value, dict_keys):
    if not isinstance(value, str):
        return None
    for key in dict_keys:
        words = key.split()
        if all(word in value for word in words):
            return key
    return None
