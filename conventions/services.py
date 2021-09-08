import datetime

from io import BytesIO
from enum import Enum
from zipfile import BadZipFile
from openpyxl import load_workbook

from conventions.models import Convention, ConventionStatut, Preteur, Pret
from programmes.models import Lot, Logement, Annexe, TypeStationnement, LogementEDD
from programmes.forms import (
    ProgrammeSelectionForm,
    ProgrammeForm,
    ProgrammmeCadastralForm,
    LogementFormSet,
    TypeStationnementFormSet,
    AnnexeFormSet,
    LogementEDDFormSet,
)
from bailleurs.forms import BailleurForm
from .forms import (
    ConventionCommentForm,
    ConventionFinancementForm,
    PretFormSet,
    UploadForm,
)

class ReturnStatus(Enum):
    SUCCESS = 'SUCCESS'
    ERROR = 'ERROR'
    WARNING = 'WARNING'

def format_date_for_form(date):
    return date.strftime("%Y-%m-%d") if date is not None else ""


def conventions_index(request, infilter):
    infilter.update(request.user.convention_filter())
    conventions = Convention.objects.prefetch_related("programme").filter(**infilter)
    return conventions


def conventions_step1(request, infilter):
    infilter.update(request.user.programme_filter())
    return (
        Lot.objects.prefetch_related("programme")
        .prefetch_related("convention_set")
        .filter(**infilter)
        .order_by("programme__nom", "financement")
    )


def select_programme_create(request):
    if request.method == "POST":
        form = ProgrammeSelectionForm(request.POST)
        if form.is_valid():
            lot = Lot.objects.get(uuid=form.cleaned_data["lot_uuid"])
            convention = Convention.objects.create(
                lot=lot,
                programme_id=lot.programme_id,
                bailleur_id=lot.bailleur_id,
                financement=lot.financement,
            )
            convention.save()
            # All is OK -> Next:
            return {
                "success": ReturnStatus.SUCCESS,
                "convention": convention,
                "form": form,
            }  # HttpResponseRedirect(reverse('conventions:step2', args=[convention.uuid]) )

    # If this is a GET (or any other method) create the default form.
    else:
        form = ProgrammeSelectionForm()

    programmes = conventions_step1(request, {})
    return {
        "success": ReturnStatus.ERROR,
        "programmes": programmes,
        "form": form,
    }  # render(request, "conventions/step1.html", {'form': form, 'programmes': programmes})


def select_programme_update(request, convention_uuid):
    convention = Convention.objects.get(uuid=convention_uuid)

    if request.method == "POST":
        #        if request.POST['convention_uuid'] is None:
        form = ProgrammeSelectionForm(request.POST)
        if form.is_valid():
            lot = Lot.objects.get(uuid=form.cleaned_data["lot_uuid"])
            convention.lot = lot
            convention.programme_id = lot.programme_id
            convention.bailleur_id = lot.bailleur_id
            convention.financement = lot.financement
            convention.save()
            # All is OK -> Next:
            return {"success": ReturnStatus.SUCCESS, "convention": convention, "form": form}

    # If this is a GET (or any other method) create the default form.
    else:
        form = ProgrammeSelectionForm(
            initial={
                "lot_uuid": str(convention.lot.uuid),
            }
        )

    programmes = conventions_step1(request, {})
    return {
        "success": ReturnStatus.ERROR,
        "programmes": programmes,
        "convention_uuid": convention_uuid,
        "form": form,
    }


def bailleur_update(request, convention_uuid):
    convention = Convention.objects.prefetch_related("bailleur").get(uuid=convention_uuid)
    bailleur = convention.bailleur

    if request.method == "POST":
        #        if request.POST['convention_uuid'] is None:
        form = BailleurForm(request.POST)
        if form.is_valid():
            bailleur.nom = form.cleaned_data["nom"]
            bailleur.siret = form.cleaned_data["siret"]
            bailleur.capital_social = form.cleaned_data["capital_social"]
            bailleur.adresse = form.cleaned_data["adresse"]
            bailleur.code_postal = form.cleaned_data["code_postal"]
            bailleur.ville = form.cleaned_data["ville"]
            bailleur.dg_nom = form.cleaned_data["dg_nom"]
            bailleur.dg_fonction = form.cleaned_data["dg_fonction"]
            bailleur.dg_date_deliberation = form.cleaned_data["dg_date_deliberation"]
            bailleur.save()
            # All is OK -> Next:
            return {"success": ReturnStatus.SUCCESS, "convention": convention, "form": form}

    # If this is a GET (or any other method) create the default form.
    else:
        form = BailleurForm(
            initial={
                "nom": bailleur.nom,
                "siret": bailleur.siret,
                "capital_social": bailleur.capital_social,
                "adresse": bailleur.adresse,
                "code_postal": bailleur.code_postal,
                "ville": bailleur.ville,
                "dg_nom": bailleur.dg_nom,
                "dg_fonction": bailleur.dg_fonction,
                "dg_date_deliberation": format_date_for_form(
                    bailleur.dg_date_deliberation
                ),
            }
        )

    return {"success": ReturnStatus.ERROR, "convention": convention, "form": form}


def programme_update(request, convention_uuid):
    convention = Convention.objects.prefetch_related("programme").get(uuid=convention_uuid)
    programme = convention.programme

    if request.method == "POST":
        #        if request.POST['convention_uuid'] is None:
        form = ProgrammeForm(request.POST)
        if form.is_valid():
            programme.adresse = form.cleaned_data["adresse"]
            programme.code_postal = form.cleaned_data["code_postal"]
            programme.ville = form.cleaned_data["ville"]
            programme.nb_logements = form.cleaned_data["nb_logements"]
            programme.type_habitat = form.cleaned_data["type_habitat"]
            programme.type_operation = form.cleaned_data["type_operation"]
            programme.anru = form.cleaned_data["anru"]
            programme.autre_locaux_hors_convention = form.cleaned_data[
                "autre_locaux_hors_convention"
            ]
            programme.nb_locaux_commerciaux = form.cleaned_data["nb_locaux_commerciaux"]
            programme.nb_bureaux = form.cleaned_data["nb_bureaux"]
            programme.save()
            # All is OK -> Next:
            return {"success": ReturnStatus.SUCCESS, "convention": convention, "form": form}

    # If this is a GET (or any other method) create the default form.
    else:
        form = ProgrammeForm(
            initial={
                "adresse": programme.adresse,
                "code_postal": programme.code_postal,
                "ville": programme.ville,
                "nb_logements": programme.nb_logements,
                "type_habitat": programme.type_habitat,
                "type_operation": programme.type_operation,
                "anru": programme.anru,
                "autre_locaux_hors_convention": programme.autre_locaux_hors_convention,
                "nb_locaux_commerciaux": programme.nb_locaux_commerciaux,
                "nb_bureaux": programme.nb_bureaux,
            }
        )

    return {"success": ReturnStatus.ERROR, "convention": convention, "form": form}

def programme_cadastral_update(request, convention_uuid):
    # pylint: disable=R0915
    convention = (Convention.objects
        .prefetch_related("programme")
        .prefetch_related("programme__logementedd_set")
        .get(uuid=convention_uuid)
    )
    programme = convention.programme

    import_warnings = None

    if request.method == "POST":
        # When the user cliked on "Téléverser" button
        formset = LogementEDDFormSet(request.POST)
        form = ProgrammmeCadastralForm(request.POST)
        if request.POST.get("Upload", False):
            upform = UploadForm(request.POST, request.FILES)
            if upform.is_valid():
                result = handle_uploaded_file(upform, request.FILES["file"], LogementEDD)
                if result['success'] != ReturnStatus.ERROR:
                    formset = LogementEDDFormSet(initial=result['objects'])
                    import_warnings = result['import_warnings']
        # When the user cliked on "Enregistrer et Suivant"
        else:
            upform = UploadForm()
            form_is_valid = form.is_valid()
            formset_is_valid = formset.is_valid()
            if form_is_valid and formset_is_valid:
                programme.permis_construire = form.cleaned_data["permis_construire"]
                programme.date_acte_notarie = form.cleaned_data["date_acte_notarie"]
                programme.date_achevement_previsible = form.cleaned_data[
                    "date_achevement_previsible"
                ]
                programme.date_achat = form.cleaned_data["date_achat"]
                programme.date_achevement = form.cleaned_data["date_achevement"]
                programme.vendeur = form.cleaned_data["vendeur"]
                programme.acquereur = form.cleaned_data["acquereur"]
                programme.reference_notaire = form.cleaned_data["reference_notaire"]
                programme.reference_publication_acte = form.cleaned_data[
                    "reference_publication_acte"
                ]
                programme.acte_de_vente = form.cleaned_data["acte_de_vente"]
                programme.edd_volumetrique = form.cleaned_data["edd_volumetrique"]
                programme.save()

                programme.logementedd_set.all().delete()
                for form_logementedd in formset:
                    logementedd = LogementEDD.objects.create(
                        programme=programme,
                        bailleur=convention.bailleur,
                        financement=form_logementedd.cleaned_data["financement"],
                        designation=form_logementedd.cleaned_data["designation"],
                        typologie=form_logementedd.cleaned_data["typologie"],
                    )
                    logementedd.save()

                # All is OK -> Next:
                return {
                    "success": ReturnStatus.SUCCESS,
                    "convention": convention,
                    "form": form,
                    "formset": formset,
                }
    # When display the file for the first time
    else:
        initial = []
        logementedds = programme.logementedd_set.all()
        for logementedd in logementedds:
            initial.append(
                {
                    "financement": logementedd.financement,
                    "designation": logementedd.designation,
                    "typologie": logementedd.typologie,
                }
            )
        formset = LogementEDDFormSet(initial=initial)
        upform = UploadForm()
        form = ProgrammmeCadastralForm(
            initial={
                "permis_construire": programme.permis_construire,
                "date_acte_notarie": format_date_for_form(programme.date_acte_notarie),
                "date_achevement_previsible": format_date_for_form(
                    programme.date_achevement_previsible
                ),
                "date_achat": format_date_for_form(programme.date_achat),
                "date_achevement": format_date_for_form(programme.date_achevement),
                "vendeur": programme.vendeur,
                "acquereur": programme.acquereur,
                "reference_notaire": programme.reference_notaire,
                "reference_publication_acte": programme.reference_publication_acte,
                "acte_de_vente": programme.acte_de_vente,
                "edd_volumetrique": programme.edd_volumetrique,
            }
        )
    return {
        "success": ReturnStatus.ERROR,
        "convention": convention,
        "form": form,
        "formset": formset,
        "upform": upform,
        "import_warnings": import_warnings,
    }


def convention_financement(request, convention_uuid):
    convention = Convention.objects.prefetch_related("pret_set").get(uuid=convention_uuid)
    import_warnings = None

    if request.method == "POST":
        # When the user cliked on "Téléverser" button
        formset = PretFormSet(request.POST)
        form = ConventionFinancementForm(request.POST)
        if request.POST.get("Upload", False):
            upform = UploadForm(request.POST, request.FILES)
            if upform.is_valid():
                result = handle_uploaded_file(upform, request.FILES["file"], Pret)
                if result['success'] != ReturnStatus.ERROR:
                    formset = PretFormSet(initial=result['objects'])
                    import_warnings = result['import_warnings']
        # When the user cliked on "Enregistrer et Suivant"
        else:
            upform = UploadForm()
            form_is_valid = form.is_valid()
            formset_is_valid = formset.is_valid()
            if form_is_valid and formset_is_valid:
                convention.date_fin_conventionnement = form.cleaned_data[
                    "date_fin_conventionnement"
                ]
                convention.fond_propre = form.cleaned_data[
                    "fond_propre"
                ]
                convention.save()
                convention.pret_set.all().delete()
                for form_pret in formset:
                    pret = Pret.objects.create(
                        convention=convention,
                        bailleur=convention.bailleur,
                        numero=form_pret.cleaned_data["numero"],
                        date_octroi=form_pret.cleaned_data["date_octroi"],
                        duree=form_pret.cleaned_data["duree"],
                        montant=form_pret.cleaned_data["montant"],
                        preteur=form_pret.cleaned_data["preteur"],
                        autre = form_pret.cleaned_data['autre'],
                    )
                    pret.save()

                # All is OK -> Next:
                return {
                    "success": ReturnStatus.SUCCESS,
                    "convention": convention,
                    "form": form,
                    "formset": formset,
                }
    # When display the file for the first time
    else:
        initial = []
        prets = convention.pret_set.all()
        for pret in prets:
            initial.append(
                {
                    "numero": pret.numero,
                    "date_octroi": format_date_for_form(pret.date_octroi),
                    "duree": pret.duree,
                    "montant": pret.montant,
                    "preteur": pret.preteur,
                    "autre": pret.autre,
                }
            )
        upform = UploadForm()
        formset = PretFormSet(initial=initial)
        form = ConventionFinancementForm(
            initial={
                "date_fin_conventionnement": format_date_for_form(
                    convention.date_fin_conventionnement
                ),
                "fond_propre": convention.fond_propre,
            }
        )
    return {
        "success": ReturnStatus.ERROR,
        "convention": convention,
        "form": form,
        "formset": formset,
        "upform": upform,
        "import_warnings": import_warnings,
    }


def logements_update(request, convention_uuid):
    convention = (
        Convention.objects
            .prefetch_related("lot")
            .prefetch_related("lot__logement_set")
            .get(uuid=convention_uuid)
    )
    import_warnings = None

    if request.method == "POST":
        # When the user cliked on "Téléverser" button
        formset = LogementFormSet(request.POST)
        if request.POST.get("Upload", False):
            upform = UploadForm(request.POST, request.FILES)
            if upform.is_valid():
                result = handle_uploaded_file(upform, request.FILES["file"], Logement)
                if result['success'] != ReturnStatus.ERROR:
                    formset = LogementFormSet(initial=result['objects'])
                    import_warnings = result['import_warnings']
        # When the user cliked on "Enregistrer et Suivant"
        else:
            upform = UploadForm()
            if formset.is_valid():
                convention.lot.logement_set.all().delete()
                for form_logement in formset:
                    logement = Logement.objects.create(
                        lot=convention.lot,
                        bailleur=convention.bailleur,
                        designation=form_logement.cleaned_data["designation"],
                        typologie=form_logement.cleaned_data["typologie"],
                        surface_habitable=form_logement.cleaned_data["surface_habitable"],
                        surface_annexes=form_logement.cleaned_data["surface_annexes"],
                        surface_annexes_retenue=
                            form_logement.cleaned_data["surface_annexes_retenue"],
                        surface_utile=form_logement.cleaned_data["surface_utile"],
                        loyer_par_metre_carre=form_logement.cleaned_data["loyer_par_metre_carre"],
                        coeficient=form_logement.cleaned_data["coeficient"],
                        loyer=form_logement.cleaned_data["loyer"],
                    )
                    logement.save()

                # All is OK -> Next:
                return {
                    "success": ReturnStatus.SUCCESS,
                    "convention": convention,
                    "formset": formset,
                }
    # When display the file for the first time
    else:
        initial = []
        logements = convention.lot.logement_set.all()
        for logement in logements:
            initial.append(
                {
                    "designation": logement.designation,
                    "typologie": logement.typologie,
                    "surface_habitable": logement.surface_habitable,
                    "surface_annexes": logement.surface_annexes,
                    "surface_annexes_retenue": logement.surface_annexes_retenue,
                    "surface_utile": logement.surface_utile,
                    "loyer_par_metre_carre": logement.loyer_par_metre_carre,
                    "coeficient": logement.coeficient,
                    "loyer": logement.loyer,
                }
            )
        upform = UploadForm()
        formset = LogementFormSet(initial=initial)
    return {
        "success": ReturnStatus.ERROR,
        "convention": convention,
        "formset": formset,
        "upform": upform,
        "import_warnings": import_warnings,
    }


def annexes_update(request, convention_uuid):
    convention = (
        Convention.objects
            .prefetch_related("lot")
            .prefetch_related("lot__logement_set")
            .prefetch_related("lot__logement_set__annexe_set")
            .get(uuid=convention_uuid)
    )
    import_warnings = None

    if request.method == "POST":
        # When the user cliked on "Téléverser" button
        formset = AnnexeFormSet(request.POST)
        if request.POST.get("Upload", False):
            upform = UploadForm(request.POST, request.FILES)
            if upform.is_valid():
                result = handle_uploaded_file(upform, request.FILES["file"], Annexe)
                if result['success'] != ReturnStatus.ERROR:
                    formset = AnnexeFormSet(initial=result['objects'])
                    import_warnings = result['import_warnings']
        # When the user cliked on "Enregistrer et Suivant"
        else:
            upform = UploadForm()
            #to do : manage this one in the model
            formset.is_valid()
            for form_annexe in formset:
                try:
                    logement = convention.lot.logement_set.get(
                        designation=form_annexe.cleaned_data["logement_designation"],
                        lot = convention.lot
                    )
                except Logement.DoesNotExist:
                    form_annexe.add_error(
                        'logement_designation',
                        "Ce logement n'existe pas dans ce lot"
                    )

            if formset.is_valid():
                Annexe.objects.filter(logement__lot_id=convention.lot.id).delete()
                for form_annexe in formset:
                    logement = Logement.objects.get(
                        designation=form_annexe.cleaned_data["logement_designation"],
                        lot = convention.lot
                    )
                    annexe = Annexe.objects.create(
                        logement=logement,
                        bailleur=convention.bailleur,
                        typologie=form_annexe.cleaned_data["typologie"],
                        surface_hors_surface_retenue=
                            form_annexe.cleaned_data["surface_hors_surface_retenue"],
                        loyer_par_metre_carre=form_annexe.cleaned_data["loyer_par_metre_carre"],
                        loyer=form_annexe.cleaned_data["loyer"],
                    )
                    annexe.save()

                # All is OK -> Next:
                return {
                    "success": ReturnStatus.SUCCESS,
                    "convention": convention,
                    "formset": formset,
                }
    # When display the file for the first time
    else:
        initial = []
        annexes = Annexe.objects.filter(logement__lot_id=convention.lot.id)
        for annexe in annexes:
            initial.append(
                {
                    "typologie": annexe.typologie,
                    "logement_designation": annexe.logement.designation,
                    "logement_typologie": annexe.logement.typologie,
                    "surface_hors_surface_retenue": annexe.surface_hors_surface_retenue,
                    "loyer_par_metre_carre": annexe.loyer_par_metre_carre,
                    "loyer": annexe.loyer,
                }
            )
        upform = UploadForm()
        formset = AnnexeFormSet(initial=initial)
    return {
        "success": ReturnStatus.ERROR,
        "convention": convention,
        "formset": formset,
        "upform": upform,
        "import_warnings": import_warnings,
    }


def stationnements_update(request, convention_uuid):
    convention = (
        Convention.objects
            .prefetch_related("lot")
            .prefetch_related("lot__typestationnement_set")
            .get(uuid=convention_uuid)
    )
    import_warnings = None

    if request.method == "POST":
        # When the user cliked on "Téléverser" button
        formset = TypeStationnementFormSet(request.POST)
        if request.POST.get("Upload", False):
            upform = UploadForm(request.POST, request.FILES)
            if upform.is_valid():
                result = handle_uploaded_file(upform, request.FILES["file"], TypeStationnement)
                if result['success'] != ReturnStatus.ERROR:
                    formset = TypeStationnementFormSet(initial=result['objects'])
                    import_warnings = result['import_warnings']
        # When the user cliked on "Enregistrer et Suivant"
        else:
            upform = UploadForm()
            if formset.is_valid():
                convention.lot.typestationnement_set.all().delete()
                for form_stationnement in formset:
                    stationnement = TypeStationnement.objects.create(
                        lot=convention.lot,
                        bailleur=convention.bailleur,
                        typologie=form_stationnement.cleaned_data["typologie"],
                        nb_stationnements=form_stationnement.cleaned_data["nb_stationnements"],
                        loyer=form_stationnement.cleaned_data["loyer"],
                    )
                    stationnement.save()

                # All is OK -> Next:
                return {
                    "success": ReturnStatus.SUCCESS,
                    "convention": convention,
                    "formset": formset,
                }
    # When display the file for the first time
    else:
        initial = []
        stationnements = convention.lot.typestationnement_set.all()
        for stationnement in stationnements:
            initial.append(
                {
                    "typologie": stationnement.typologie,
                    "nb_stationnements": stationnement.nb_stationnements,
                    "loyer": stationnement.loyer,
                }
            )
        upform = UploadForm()
        formset = TypeStationnementFormSet(initial=initial)
    return {
        "success": ReturnStatus.ERROR,
        "convention": convention,
        "formset": formset,
        "upform": upform,
        "import_warnings": import_warnings,
    }


def convention_comments(request, convention_uuid):
    convention = Convention.objects.get(uuid=convention_uuid)

    if request.method == "POST":
        form = ConventionCommentForm(request.POST)
        if form.is_valid():
            convention.comments = form.cleaned_data["comments"]
            convention.save()
            # All is OK -> Next:
            return {"success": ReturnStatus.SUCCESS, "convention": convention, "form": form}

    else:
        form = ConventionCommentForm(
            initial={
                "comments": convention.comments,
            }
        )

    return {"success": ReturnStatus.ERROR, "convention": convention, "form": form}


def convention_summary(request, convention_uuid):
    convention = (
        Convention.objects
            .prefetch_related("bailleur")
            .prefetch_related("programme")
            .prefetch_related("lot")
            .prefetch_related("lot__typestationnement_set")
            .prefetch_related("lot__logement_set")
            .get(uuid=convention_uuid)
    )
    if request.method == "POST":
        if request.POST.get("GenerateConvention", False):
            print("GenerateConvention")
            return {
                "success": ReturnStatus.ERROR,
                "convention": convention,
                "bailleur": convention.bailleur,
                "lot": convention.lot,
                "programme": convention.programme,
                "logements": convention.lot.logement_set.all(),
                "stationnements": convention.lot.typestationnement_set.all(),
                "annexes": Annexe.objects.filter(logement__lot_id=convention.lot.id).all(),
            }
        if request.POST.get("SubmitConvention", False):
            convention.soumis_le = datetime.date.today()
            convention.statut = ConventionStatut.INSTRUCTION
            convention.save()
            return {
                "success": ReturnStatus.SUCCESS,
                "convention": convention,
            }
        return {
            "success": ReturnStatus.WARNING,
            "convention": convention,
        }
    return {
        "success": ReturnStatus.ERROR,
        "convention": convention,
        "bailleur": convention.bailleur,
        "lot": convention.lot,
        "programme": convention.programme,
        "logements": convention.lot.logement_set.all(),
        "stationnements": convention.lot.typestationnement_set.all(),
        "annexes": Annexe.objects.filter(logement__lot_id=convention.lot.id).all(),
    }


def handle_uploaded_file(upform, my_file, myClass):
    import_mapping = myClass.import_mapping
    try:
        my_wb = load_workbook(filename=BytesIO(my_file.read()), data_only=True)
    except BadZipFile:
        upform.add_error(
            'file',
            "Le fichier importé ne semble pas être du bon format, 'xlsx' est le format attendu"
        )
        return {"success": ReturnStatus.ERROR}
    try:
        my_ws = my_wb[myClass.sheet_name]
    except KeyError:
        upform.add_error(
            'file',
            f"Le fichier importé doit avoir une feuille nommée '{myClass.sheet_name}'"
        )
        return {"success": ReturnStatus.ERROR}
    import_warnings = []
    column_from_index = {}
    for col in my_ws.iter_cols(min_col=1, max_col=my_ws.max_column, min_row=1, max_row=1):
        for cell in col:
            if cell.value is None:
                continue
            if cell.value not in import_mapping:
                import_warnings.append(Exception(
                    f"La colonne nommée '{cell.value}' est inconnue, elle sera ignorée. " +
                    f"Les colonnes attendues sont : {', '.join(import_mapping.keys())}"
                ))
                continue
            column_from_index[cell.column] = str(cell.value).strip()

    # transform each line into object
    my_objects = []
    for row in my_ws.iter_rows(
        min_row=3, max_row=my_ws.max_row, min_col=1, max_col=my_ws.max_column
    ):
        my_row, empty_line, new_warnings = extract_row(row, column_from_index, import_mapping)
        import_warnings = [*import_warnings, *new_warnings]

        # Ignore if the line is empty
        if not empty_line:
            my_objects.append(my_row)

    return {
        "success": ReturnStatus.SUCCESS if len(import_warnings) == 0 else ReturnStatus.WARNING,
        'objects': my_objects,
        "import_warnings": import_warnings,
    }


def extract_row(row, column_from_index, import_mapping):
    # pylint: disable=R0912
    new_warnings = []
    my_row = {}
    empty_line = True
    for cell in row:
        # Ignore unknown column
        if cell.column not in column_from_index:
            continue

        # Check the empty lines to don't fill it
        if cell.value is not None:
            empty_line = False
        else:
            continue


        value = None
        model_field = import_mapping[column_from_index[cell.column]]

        if isinstance(model_field, str):
            key = model_field
            value = cell.value
        else:
            key = model_field.name

            # Date case
            if model_field.get_internal_type() == 'DateField':
                if isinstance(cell.value, datetime.datetime):
                    value = format_date_for_form(cell.value)
                else:
                    new_warnings.append(Exception(
                        f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' " +
                        f"de la colonne {column_from_index[cell.column]} " +
                        "doit être une date"
                    ))

            # TextChoices case
            elif (model_field.get_internal_type() == 'CharField' and
                    model_field.choices is not None):
                if cell.value is not None:
                    value = next((x[0] for x in model_field.choices if x[1] == cell.value), None)
                    if value is None: # value is not Null but not in the choices neither
                        new_warnings.append(Exception(
                            f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' " +
                            f"de la colonne {column_from_index[cell.column]} " +
                            f"doit faire partie des valeurs : {', '.join(Preteur.labels)}"
                        ))

            # Float case
            elif model_field.get_internal_type() == 'FloatField':
                if cell.value is not None:
                    if isinstance(cell.value, (float, int)):
                        value = float(cell.value)
                    else:
                        new_warnings.append(Exception(
                            f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' " +
                            f"de la colonne {column_from_index[cell.column]} " +
                            "doit être une valeur numérique"
                        ))

            # Integer case
            elif model_field.get_internal_type() == 'IntegerField':
                if cell.value is not None:
                    if isinstance(cell.value, (float, int)):
                        value = int(cell.value)
                    else:
                        new_warnings.append(Exception(
                            f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' " +
                            f"de la colonne {column_from_index[cell.column]} " +
                            "doit être une valeur numérique"
                        ))

            # String case
            elif model_field.get_internal_type() == 'CharField':
                if cell.value is not None:
                    if isinstance(cell.value, (float, int, str)):
                        value = cell.value
                    else:
                        new_warnings.append(Exception(
                            f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' " +
                            f"de la colonne {column_from_index[cell.column]} " +
                            "doit être une valeur alphanumeric"
                        ))
        my_row[key] = value

    return my_row, empty_line, new_warnings
