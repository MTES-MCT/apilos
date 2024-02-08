# ce script ne marche pas : incompatibilité de openpyyxl et csv

from openpyxl import load_workbook

wb2 = load_workbook("../documents/export_sisal.csv")
print(wb2.sheetnames)  # noqa: T201
