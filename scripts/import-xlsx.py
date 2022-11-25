from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
wb = load_workbook('../documents/export_sisal.xlsx')

sheet_name = 'Rapport 1'
ws = wb[sheet_name]
first_cell_of_table = 'B4'

# Create one object by row
column_from_index = {}
for tuple in ws['B4':'AH4']:
  for cell in tuple:
    column_from_index[cell.column] = cell.value
    print(cell.value)

my_objects = []
for row in ws.iter_rows(min_row=5,max_row=ws.max_row,min_col=2, max_col=ws.max_column): #(    'B{}:AH{}'.format(4,int(ws.max_row))):
  my_row = {}
  for cell in row:
    my_row[column_from_index[cell.column]] = cell.value
  my_objects.append(my_row)

#print ("\n==== NEXT ====")
#print(my_objects)



# Close the workbook after reading
wb.close()
